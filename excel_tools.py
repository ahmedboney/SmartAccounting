# -*- coding: utf-8 -*-
"""أدوات الإكسيل: تصدير منسّق + استيراد من القوالب"""
import io
import re
from datetime import datetime, date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.page import PageMargins

# ---------- الألوان (رسمي: خط أسود Bold وحدود سوداء — مناسب للطباعة) ----------
INK = "000000"        # كل الخطوط أسود
HDR_GRAY = "D9D9D9"   # رمادي فاتح لرأس الجدول
ALT = "F2F2F2"        # تظليل الصفوف الفردية
TOT_GRAY = "BFBFBF"   # صف الإجمالي

# أسماء قديمة للحفاظ على توافق الكود الداخلي
NAVY = HDR_GRAY
NAVY_DARK = HDR_GRAY
GOLD = TOT_GRAY
BORDER_C = INK

THIN = Side(style="thin", color=BORDER_C)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

F_TITLE = Font(name="Arial", size=16, bold=True, color=INK)
F_SUB = Font(name="Arial", size=11, bold=True, color=INK)
F_HEADER = Font(name="Arial", size=11, bold=True, color=INK)
F_DATA = Font(name="Arial", size=11, bold=True, color=INK)
F_TOTAL = Font(name="Arial", size=12, bold=True, color=INK)
F_SIGN = Font(name="Arial", size=12, bold=True, color=INK)

FILL_TITLE = PatternFill("solid", fgColor=HDR_GRAY)
FILL_HEADER = PatternFill("solid", fgColor=HDR_GRAY)
FILL_ALT = PatternFill("solid", fgColor=ALT)
FILL_TOTAL = PatternFill("solid", fgColor=TOT_GRAY)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT_NUM = Alignment(horizontal="center", vertical="center")

SIGN_TITLES = ["المحاسب"]


def _auto_widths(ws, headers, rows, min_w=10, max_w=44):
    for i in range(len(headers)):
        col = get_column_letter(i + 1)
        longest = len(str(headers[i]))
        for r in rows:
            v = r[i] if i < len(r) else ""
            if isinstance(v, (datetime, date)):
                longest = max(longest, 10)
            else:
                longest = max(longest, len(str(v if v is not None else "")))
        ws.column_dimensions[col].width = min(max(min_w, longest * 1.35 + 4), max_w)


def build_report(sheet_name, title, subtitles, headers, rows,
                 totals=None, num_cols=(), date_cols=(), landscape=True):
    """يبني ورقة تقرير كاملة التنسيق ويرجع (wb, ws, next_row)"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.rightToLeft = True
    ncol = len(headers)

    # شريط العنوان
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value=title)
    c.font = F_TITLE
    c.fill = FILL_TITLE
    c.alignment = CENTER
    ws.row_dimensions[1].height = 32

    # أسطر البيانات الفرعية (الشركة / الفترة / المنطقة ...)
    r = 2
    for sub in subtitles:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        c = ws.cell(row=r, column=1, value=sub)
        c.font = F_SUB
        c.fill = FILL_TITLE
        c.alignment = CENTER
        ws.row_dimensions[r].height = 18
        r += 1
    r += 1  # صف فاصل

    # رأس الجدول
    hdr_row = r
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=i, value=h)
        c.font = F_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER_ALL
    ws.row_dimensions[hdr_row].height = 26

    # البيانات
    rr = hdr_row + 1
    for idx, row in enumerate(rows):
        for i in range(ncol):
            v = row[i] if i < len(row) else None
            c = ws.cell(row=rr, column=i + 1, value=v)
            c.font = F_DATA
            c.border = BORDER_ALL
            c.alignment = LEFT_NUM if (i in num_cols or i in date_cols) else RIGHT
            if idx % 2 == 1:
                c.fill = FILL_ALT
            if i in num_cols:
                c.number_format = "#,##0.00"
            elif i in date_cols:
                c.number_format = "dd/mm/yyyy"
        rr += 1

    # صف الإجمالي
    if totals is not None and len(rows) > 0:
        for i in range(ncol):
            v = totals[i] if i < len(totals) else None
            c = ws.cell(row=rr, column=i + 1, value=v)
            c.font = F_TOTAL
            c.fill = FILL_TOTAL
            c.border = BORDER_ALL
            c.alignment = LEFT_NUM if i in num_cols else CENTER
            if i in num_cols:
                c.number_format = "#,##0.00"
        ws.row_dimensions[rr].height = 22
        rr += 1

    _auto_widths(ws, headers, rows)

    # تجميد + فلترة + طباعة
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(ncol)}{max(hdr_row, rr - 1)}"
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    # تنسيقات الطباعة: تكرار رأس الجدول في كل صفحة + هوامش + ترقيم صفحات
    ws.print_title_rows = f"{hdr_row}:{hdr_row}"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6,
                                  header=0.3, footer=0.3)
    ws.oddFooter.center.text = "صفحة &P من &N"
    ws.oddFooter.center.size = 9
    ws.oddFooter.center.font = "Arial"

    return wb, ws, rr


def _sig_positions(ncol, count):
    """مواضع متباعدة غير متداخلة داخل عدد الأعمدة حسب عدد التواقيع"""
    if ncol < 1 or count < 1:
        return []
    if count >= ncol:
        return list(range(1, ncol + 1))
    positions = set()
    for i in range(count):
        p = round(ncol * (i + 1) / (count + 1))
        positions.add(max(1, min(ncol, p)))
    return sorted(positions)


def _signature_brackets(name):
    """يولّد سطر التوقيع بأقواس عرضها مناسب لطول الاسم اللي تحته"""
    inner = max(14, min(50, int(len(name) * 1.9)))
    return f"التوقيع ({' ' * inner})"


def add_signatures(ws, row, ncol, acc_user="", signatures=None):
    """كتلة التواقيع قابلة للتعديل من صفحة الإعدادات:
    أول توقيع للمحاسب (اسم مستخدم السيستم ديناميكي)
    + قائمة تواقيع إضافية [{title, name}] محفوظة في الإعدادات.
    """
    extras = signatures or []
    row = row + 2
    titles = [SIGN_TITLES[0]]
    names = [f"محاسب / {acc_user}" if acc_user else "محاسب"]
    for s in extras:
        t = (s.get("title") or "").strip()
        n = (s.get("name") or "").strip()
        if t and n:
            titles.append(t)
            names.append(n)
    for title, name, p in zip(titles, names, _sig_positions(ncol, len(titles))):
        c0 = ws.cell(row=row, column=p, value=_signature_brackets(name))
        c0.font = F_SIGN
        c0.alignment = CENTER
        c1 = ws.cell(row=row + 1, column=p, value=name)
        c1.font = Font(name="Arial", size=11, bold=True, color=INK)
        c1.alignment = CENTER
        c2 = ws.cell(row=row + 2, column=p, value=title)
        c2.font = Font(name="Arial", size=11, bold=True, color=INK)
        c2.alignment = CENTER
    ws.row_dimensions[row].height = 38
    ws.row_dimensions[row + 1].height = 18
    ws.row_dimensions[row + 2].height = 20
    return row + 2


def to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def fmt_date(d):
    if isinstance(d, (datetime, date)):
        return d.strftime("%d/%m/%Y")
    return str(d or "")


# ======================================================================
# الاستيراد
# ======================================================================

AR_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹", "01234567890123456789")


def norm_text(v):
    if v is None:
        return ""
    s = str(v).translate(AR_DIGITS)
    return re.sub(r"\s+", " ", s).strip()


def parse_number(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = norm_text(v).replace(",", "").replace("٫", ".").replace("٬", "")
    if s in ("-", ""):
        return 0.0
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    try:
        n = float(s)
        return -n if neg else n
    except ValueError:
        return 0.0


def parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = norm_text(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def find_header(ws, synonyms):
    """يبحث عن صف العناوين ويرجع خريطة: المفتاح -> رقم العمود"""
    best_row, best_map, best_score = None, {}, 0
    max_r = min(ws.max_row, 30)
    max_c = min(ws.max_column, 40)
    for r in range(1, max_r + 1):
        cells = {}
        for c in range(1, max_c + 1):
            t = norm_text(ws.cell(row=r, column=c).value)
            if t:
                cells[c] = t
        if len(cells) < 2:
            continue
        mapping, score = {}, 0
        for key, keys_list in synonyms.items():
            found = None
            for c, t in cells.items():
                tl = t.replace("إ", "ا").replace("أ", "ا").replace("ى", "ي").replace("ة", "ه")
                for k in keys_list:
                    k2 = k.replace("إ", "ا").replace("أ", "ا").replace("ى", "ي").replace("ة", "ه")
                    if t == k2 or tl == k2:
                        found = c
                        break
                if found:
                    break
            if not found:
                for c, t in cells.items():
                    for k in keys_list:
                        if k in t:
                            found = c
                            break
                    if found:
                        break
            if found:
                mapping[key] = found
                score += 1
        if score > best_score:
            best_score, best_row, best_map = score, r, mapping
    return best_row, best_map, best_score


SYN_ACCOUNTS = {
    "name": ["اسم الحساب", "اسم الحساب", "الحساب", "الاسم"],
    "acc_no": ["رقم الحساب", "كود الحساب", "الكود", "رقم"],
    "type": ["النوع", "نوع الحساب"],
    "opening": ["رصيد افتتاحي", "افتتاحي"],
}

SYN_JOURNAL = {
    "movement": ["رقم الحركة", "رقم الحركه", "رقم القيد", "قيد", "الحركة"],
    "date": ["التاريخ", "تاريخ"],
    "debit": ["مدين", "له"],
    "credit": ["دائن", "عليه"],
    "name": ["اسم الحساب", "إسم الحساب", "اسم الحساب", "الحساب", "الاسم"],
    "acc_no": ["رقم الحساب", "كود الحساب", "الكود"],
    "desc": ["البيان", "شرح", "الوصف", "ملاحظات"],
    "region": ["المنطقة", "المنطقه", "منطقة", "الفرع", "فرع"],
}


def read_accounts_template(stream):
    """يقرأ دليل حسابات.xlsx ويرجع (قائمة صفوف, أخطاء)"""
    wb = load_workbook(stream, data_only=True)
    ws = wb.active
    errors, out = [], []
    hdr_row, hmap, score = find_header(ws, SYN_ACCOUNTS)
    if not hdr_row or "acc_no" not in hmap or "name" not in hmap:
        return [], ["لم يتم العثور على عمودي (اسم الحساب / رقم الحساب). تأكد من استخدام القالب الصحيح."]
    for r in range(hdr_row + 1, ws.max_row + 1):
        acc_no = norm_text(ws.cell(row=r, column=hmap["acc_no"]).value)
        name = norm_text(ws.cell(row=r, column=hmap["name"]).value)
        acc_type = norm_text(ws.cell(row=r, column=hmap["type"]).value) if "type" in hmap else ""
        opening = parse_number(ws.cell(row=r, column=hmap["opening"]).value) if "opening" in hmap else 0.0
        if not acc_no and not name:
            continue
        if not acc_no or not name:
            errors.append(f"صف {r}: اسم الحساب ورقم الحساب مطلوبان معًا — تم تخطيه")
            continue
        out.append({"acc_no": acc_no, "name": name, "type": acc_type, "opening": opening})
    return out, errors


def read_journal_template(stream):
    """يقرأ دفتر اليومية.xlsx ويرجع (مجموعات القيود, أخطاء)"""
    wb = load_workbook(stream, data_only=True)
    ws = wb.active
    errors, groups, order = [], {}, []
    hdr_row, hmap, score = find_header(ws, SYN_JOURNAL)
    required = {"movement", "acc_no"}
    if not hdr_row or not required.issubset(hmap.keys()):
        return [], ["لم يتم التعرف على أعمدة القالب. تأكد أن الملف يحتوي: رقم الحركة، التاريخ، مدين، دائن، إسم الحساب، رقم الحساب، البيان."]

    def cellv(r, key):
        if key not in hmap:
            return None
        return ws.cell(row=r, column=hmap[key]).value

    auto_seq = 0
    for r in range(hdr_row + 1, ws.max_row + 1):
        acc_no = norm_text(cellv(r, "acc_no"))
        name = norm_text(cellv(r, "name"))
        debit = parse_number(cellv(r, "debit"))
        credit = parse_number(cellv(r, "credit"))
        if not acc_no and debit == 0 and credit == 0:
            continue
        mv_raw = norm_text(cellv(r, "movement"))
        if not mv_raw:
            auto_seq += 1
            mv_raw = f"#تلقائي-{auto_seq}"
        dt = parse_date(cellv(r, "date"))
        desc = norm_text(cellv(r, "desc"))
        region = norm_text(cellv(r, "region"))
        line = {"row": r, "acc_no": acc_no, "name": name, "debit": debit, "credit": credit}
        if debit != 0 and credit != 0:
            errors.append(f"صف {r}: لا يمكن إدخال مدين ودائن معًا — تم تخطيه")
            continue
        if debit == 0 and credit == 0:
            errors.append(f"صف {r}: قيمة المدين والدائن صفر — تم تخطيه")
            continue
        if mv_raw not in groups:
            groups[mv_raw] = {"movement": mv_raw, "date": dt, "desc": desc,
                              "region": region, "lines": []}
            order.append(mv_raw)
        g = groups[mv_raw]
        g["lines"].append(line)
        if dt and not g["date"]:
            g["date"] = dt
        if desc and not g["desc"]:
            g["desc"] = desc
        if region and not g["region"]:
            g["region"] = region
    return [groups[k] for k in order], errors


# ======================================================================
# تصديرات التقارير
# ======================================================================

def export_accounts(accounts, company, acc_user="", signatures=None):
    headers = ["رقم الحساب", "اسم الحساب", "النوع", "الرصيد الافتتاحي"]
    rows = [[a["acc_no"], a["name"], a["type"] or "", a["opening_balance"]] for a in accounts]
    total = sum(a["opening_balance"] for a in accounts)
    wb, ws, nr = build_report(
        "دليل حسابات", "دليل الحسابات",
        [company],
        headers, rows,
        totals=["", "الإجمالي", "", total],
        num_cols=(3,), landscape=False,
    )
    add_signatures(ws, nr, len(headers), acc_user=acc_user, signatures=signatures)
    return to_bytes(wb)


def export_journal(entries, company, subtitle_extra="", acc_user="", signatures=None):
    headers = ["رقم الحركة", "التاريخ", "رقم الحساب", "إسم الحساب", "البيان",
               "المنطقة", "نوع القيد", "الحالة", "مدين", "دائن"]
    rows, td, tc = [], 0.0, 0.0
    for e in entries:
        status = "مرحل" if e["status"] == "posted" else "مسودة"
        etype = e.get("entry_type") or "عادي"
        d = datetime.strptime(e["entry_date"], "%Y-%m-%d").date() if isinstance(e["entry_date"], str) else e["entry_date"]
        for ln in e["lines"]:
            rows.append([e["movement_no"], d, ln["acc_no"], ln["name"],
                         e["description"], e["region"], etype, status, ln["debit"], ln["credit"]])
            td += ln["debit"]
            tc += ln["credit"]
    subs = [company]
    if subtitle_extra:
        subs.append(subtitle_extra)
    wb, ws, nr = build_report(
        "دفتر اليومية", "دفتر اليومية", subs, headers, rows,
        totals=["", "", "", "", "", "", "", "الإجمالي", round(td, 2), round(tc, 2)],
        num_cols=(8, 9), date_cols=(1,),
    )
    add_signatures(ws, nr, len(headers), acc_user=acc_user, signatures=signatures)
    return to_bytes(wb)


def export_ledger(account, opening, rows, totals, company, subtitle_extra="", acc_user="", signatures=None):
    headers = ["التاريخ", "رقم الحركة", "البيان", "المنطقة", "مدين", "دائن", "الرصيد"]
    xrows = []
    for rw in rows:
        d = datetime.strptime(rw["date"], "%Y-%m-%d").date() if isinstance(rw["date"], str) else rw["date"]
        xrows.append([d, rw["movement_no"], rw["description"], rw["region"],
                      rw["debit"], rw["credit"], rw["balance"]])
    subs = [company, f"الحساب: {account['acc_no']} - {account['name']}", f"الرصيد الافتتاحي: {opening:,.2f}"]
    if subtitle_extra:
        subs.append(subtitle_extra)
    wb, ws, nr = build_report(
        "الأستاذ العام", "الأستاذ العام", subs, headers, xrows,
        totals=["", "", "الإجمالي", "",
                round(totals["debit"], 2), round(totals["credit"], 2), round(totals["balance"], 2)],
        num_cols=(4, 5, 6), date_cols=(0,),
    )
    add_signatures(ws, nr, len(headers), acc_user=acc_user, signatures=signatures)
    return to_bytes(wb)


def export_trial(rows, totals, company, subtitle_extra="", acc_user="", signatures=None):
    headers = ["رقم الحساب", "اسم الحساب", "افتتاحي مدين", "افتتاحي دائن",
               "حركة مدين", "حركة دائن", "رصيد مدين", "رصيد دائن"]
    subs = [company]
    if subtitle_extra:
        subs.append(subtitle_extra)
    wb, ws, nr = build_report(
        "ميزان المراجعة", "ميزان المراجعة", subs, headers, rows,
        totals=["", "الإجمالي"] + [round(totals[k], 2) for k in
                                   ["o_debit", "o_credit", "p_debit", "p_credit", "f_debit", "f_credit"]],
        num_cols=(2, 3, 4, 5, 6, 7),
    )
    add_signatures(ws, nr, len(headers), acc_user=acc_user, signatures=signatures)
    return to_bytes(wb)


def export_budget(entries, summary, company, subtitle_extra="", acc_user="", signatures=None):
    headers = ["رقم الحساب", "اسم الحساب", "النوع", "الميزانية", "المنفّذ", "الفرق", "النسبة %", "الحالة"]
    rows = []
    for e in entries:
        rows.append([e["acc_no"], e["name"], e["type"], e["budget"],
                      e["actual"], e["variance"], e["pct"], e["status"]])
    subs = [company]
    if subtitle_extra:
        subs.append(subtitle_extra)
    wb, ws, nr = build_report(
        "الموازنة التقديرية", "تقرير الموازنة — Budget vs Actual", subs, headers, rows,
        totals=["", "الإجمالي", "", round(summary["budget"], 2),
                round(summary["actual"], 2), round(summary["variance"], 2),
                round(summary["pct"], 1), ""],
        num_cols=(3, 4, 5, 6),
    )
    for row in range(2, nr + 1):
        status_cell = ws.cell(row=row, column=8)
        val = str(status_cell.value or "")
        if "تجاوز" in val:
            status_cell.font = Font(bold=True, color="FF0000")
        elif "قريب" in val:
            status_cell.font = Font(bold=True, color="FF8C00")
        elif "ضمن" in val:
            status_cell.font = Font(bold=True, color="008000")
        var_cell = ws.cell(row=row, column=6)
        if isinstance(var_cell.value, (int, float)) and var_cell.value < 0:
            var_cell.font = Font(bold=True, color="FF0000")
    add_signatures(ws, nr, len(headers), acc_user=acc_user, signatures=signatures)
    return to_bytes(wb)


# ======================================================================
# قوالب فارغة للتحميل
# ======================================================================

def _style_template_rows(ws, ncol, nrows, num_cols=(), date_cols=()):
    """حدود سوداء وتنسيقات جاهزة على صفوف البيانات الفارغة عشان الملء يبقى مرتب"""
    for r in range(2, 2 + nrows):
        for c in range(1, ncol + 1):
            ws.cell(row=r, column=c).border = BORDER_ALL
        for c in num_cols:
            ws.cell(row=r, column=c).number_format = "#,##0.00"
        for c in date_cols:
            ws.cell(row=r, column=c).number_format = "dd/mm/yyyy"


def blank_template_accounts():
    """قالب دليل الحسابات — مطابق تمامًا لملف المستخدم الأصلي + حدود جاهزة للملء"""
    wb = Workbook()
    ws = wb.active
    ws.title = "دليل حسابات"
    ws.sheet_view.rightToLeft = True
    headers = ["اسم الحساب", "رقم الحساب"]
    widths = [36, 16]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = F_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER_ALL
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[1].height = 24
    _style_template_rows(ws, len(headers), nrows=200)
    ws.freeze_panes = "A2"
    return to_bytes(wb)


def blank_template_journal():
    """قالب دفتر اليومية — مطابق تمامًا لملف المستخدم الأصلي + حدود جاهزة للملء"""
    wb = Workbook()
    ws = wb.active
    ws.title = "دفتر اليومية"
    ws.sheet_view.rightToLeft = True
    headers = ["رقم الحركة", "التاريخ", "مدين", "دائن", "إسم الحساب", "رقم الحساب", "البيان", "المنطقة"]
    widths = [13, 13, 14, 14, 32, 14, 42, 20]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = F_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER_ALL
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[1].height = 24
    # مدين=عمود3 ، دائن=عمود4 بأرقام ، التاريخ=عمود2 بتاريخ
    _style_template_rows(ws, len(headers), nrows=400, num_cols=(3, 4), date_cols=(2,))
    ws.freeze_panes = "A2"
    return to_bytes(wb)
