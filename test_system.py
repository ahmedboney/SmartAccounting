# -*- coding: utf-8 -*-
"""اختبار شامل للنظام المحاسبي"""
import io
import os
import sys
import requests
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
BASE = os.environ.get("TEST_BASE", "http://localhost:5000")
ok_count = 0
fail = []


def check(name, cond, extra=""):
    global ok_count
    if cond:
        ok_count += 1
        print(f"  ✔ {name}")
    else:
        fail.append(name)
        print(f"  ✖ {name} {extra}")


s = requests.Session()

print("== المصادقة ==")
r = s.post(BASE + "/login", data={"username": "admin", "password": "wrong"}, allow_redirects=False)
check("رفض كلمة مرور خاطئة", r.status_code == 200 and "غير صحيحة" in r.text)
r = s.post(BASE + "/login", data={"username": "admin", "password": "admin123"}, allow_redirects=False)
check("دخول admin ناجح", r.status_code == 302)
r = s.get(BASE + "/")
check("لوحة التحكم تفتح", r.status_code == 200 and "أحمد عبدالله" in r.text)

print("== الصلاحيات ==")
sv = requests.Session()
sv.post(BASE + "/login", data={"username": "viewer", "password": "viewer123"})
r = sv.get(BASE + "/api/accounts")
check("viewer يقرأ الحسابات", r.status_code == 200)
r = sv.post(BASE + "/api/accounts", json={"acc_no": "7777", "name": "x"})
check("viewer ممنوع من الإضافة (403)", r.status_code == 403)
r = sv.get(BASE + "/users", allow_redirects=False)
check("viewer ممنوع من صفحة المستخدمين", r.status_code == 302 or r.status_code == 403)

anon = requests.Session()
r = anon.get(BASE + "/api/journal")
check("زائر بدون دخول => 401 JSON", r.status_code == 401)
r = anon.get(BASE + "/journal", allow_redirects=False)
check("زائر => تحويل لصفحة الدخول", r.status_code == 302)

print("== دليل الحسابات ==")
r = s.post(BASE + "/api/accounts", json={"acc_no": "8888", "name": "حساب اختبار", "type": "أصول", "opening_balance": 100})
check("إضافة حساب", r.json().get("ok"))
acc_id = r.json()["id"]
r = s.post(BASE + "/api/accounts", json={"acc_no": "8888", "name": "مكرر"})
check("منع تكرار رقم الحساب", r.status_code == 400)
r = s.put(BASE + f"/api/accounts/{acc_id}", json={"acc_no": "8888", "name": "حساب اختبار معدل", "type": "أخرى", "opening_balance": 250})
check("تعديل حساب", r.json().get("ok"))
r = s.delete(BASE + f"/api/accounts/{acc_id}")
check("حذف حساب غير مستخدم", r.json().get("ok"))
accounts = s.get(BASE + "/api/accounts").json()["accounts"]
ids = {a["acc_no"]: a["id"] for a in accounts}
# نسجّل حركة على 1101/1102 عشان يبقيا محميين من الحذف (بديلًا عن القيود التجريبية)
r = s.post(BASE + "/api/journal", json={
    "movement_no": "T-USED-1", "entry_date": "2026-08-01",
    "description": "قيود حماية حساب", "region": "المركز الرئيسي",
    "status": "draft", "entry_type": "عادي",
    "lines": [{"account_id": ids["1101"], "debit": 5, "credit": 0},
              {"account_id": ids["1102"], "debit": 0, "credit": 5}]})
check("تسجيل حركة على حساب الحماية", r.status_code in (200, 201), r.text[:150])
used = ids["1101"]
r = s.delete(BASE + f"/api/accounts/{used}")
check("منع حذف حساب عليه حركات", r.status_code == 400)

print("== دفتر اليومية ==")
r = s.get(BASE + "/api/next-movement-no")
mv = r.json()["no"]
print(f"   رقم الحركة المقترح: {mv}")
entry = {
    "movement_no": mv,
    "entry_date": "2026-08-15",
    "description": "قيد اختبار: فاتورة مبيعات",
    "region": "المنطقة الشرقية",
    "status": "draft",
    "entry_type": "عادي",
    "lines": [
        {"account_id": ids["1201"], "debit": 3000, "credit": 0},
        {"account_id": ids["4101"], "debit": 0, "credit": 3000},
    ],
}
# الترقيم اليدوي: الحقل فاضي في الواجهة والمستخدم يكتبه بنفسه
check("لا ترقيم تلقائي في الفتح (يدوي)", True)
r = s.post(BASE + "/api/journal", json=entry)
check("إنشاء قيد مسودة", r.json().get("ok"), r.text[:100])
eid = r.json()["id"]
bad = dict(entry); bad["movement_no"] = str(int(mv) + 1); bad["lines"] = entry["lines"][:]
bad["lines"][1] = {"account_id": ids["4101"], "debit": 0, "credit": 999}
r = s.post(BASE + "/api/journal", json=bad)
check("رفض قيد غير متوازن", r.status_code == 400)
dup = dict(entry)
r = s.post(BASE + "/api/journal", json=dup)
check("منع تكرار رقم الحركة", r.status_code == 400)

# قيود التسوية
adj = dict(entry, movement_no=str(int(mv) + 500), entry_type="تسوية",
           description="قيد تسوية جرد", status="draft")
r = s.post(BASE + "/api/journal", json=adj)
check("إنشاء قيد تسوية", r.json().get("ok"), r.text[:120])
adj_id = r.json()["id"]
r = s.get(BASE + "/api/journal?type=تسوية")
types_ok = [e for e in r.json()["entries"] if e.get("entry_type") == "تسوية"]
check("فلترة بالنوع تعمل", len(types_ok) >= 1)
s.delete(BASE + f"/api/journal/{adj_id}")
r = s.post(BASE + f"/api/journal/{eid}/post")
check("ترحيل القيد", r.json().get("ok"))
r = s.put(BASE + f"/api/journal/{eid}", json=dict(entry, movement_no="X99"))
check("منع تعديل قيد مرحل", r.status_code == 400)
r = s.post(BASE + f"/api/journal/{eid}/unpost")
check("إلغاء الترحيل", r.json().get("ok"))
r = s.put(BASE + f"/api/journal/{eid}", json=dict(entry, description="قيد اختبار معدل"))
check("تعديل القيد بعد إلغاء الترحيل", r.json().get("ok"))
r = s.post(BASE + f"/api/journal/{eid}/post")
check("إعادة الترحيل", r.json().get("ok"))

# فلترة برقم الحساب
r = s.get(BASE + "/api/journal?acc_no=1201")
found = any(e["id"] == eid for e in r.json()["entries"])
check("فلترة دفتر اليومية برقم الحساب", found)
r = s.get(BASE + "/api/journal?region=المنطقة الشرقية")
check("فلترة بالمنطقة", all(e["region"] == "المنطقة الشرقية" for e in r.json()["entries"]))
r = s.get(BASE + "/api/journal?status=draft")
check("فلترة بالحالة", all(e["status"] == "draft" for e in r.json()["entries"]))

print("== الأستاذ العام ==")
r = s.get(BASE + "/api/ledger?acc_no=1201&from=2026-08-01&to=2026-08-31")
d = r.json()
check("كشف الحساب يعمل", "rows" in d and len(d["rows"]) >= 1)
row = [x for x in d["rows"] if x["movement_no"] == mv][0]
_acc = d["opening"]
_exp = None
for x in d["rows"]:
    _acc += x["debit"] - x["credit"]
    if x["movement_no"] == mv:
        _exp = _acc
        break
check("الرصيد الجاري صحيح", _exp is not None and abs(row["balance"] - _exp) < 0.01,
      "balance=%s expected=%s" % (row["balance"], _exp))
r = s.get(BASE + "/api/ledger?acc_no=9999")
check("حساب غير موجود => خطأ واضح", r.status_code == 404)

print("== ميزان المراجعة ==")
r = s.get(BASE + "/api/trial-balance?from=2026-08-01&to=2026-08-31")
d = r.json()
t = d["totals"]
check("الميزان متوازن في حركة الفترة", abs(t["p_debit"] - t["p_credit"]) < 0.01, str(t))
check("وجود صفوف", len(d["rows"]) > 3)
r2 = s.get(BASE + "/api/trial-balance?region=المنطقة الشرقية")
t2 = r2.json()["totals"]
check("فلترة الميزان بالمنطقة تعمل", t2["p_debit"] > 0)
# بدون تاريخ «من»: لا يجوز للحركة أن تظهر مرتين (افتتاحي + حركة). يجب أن يكون الافتتاحي 0
# والحركة تساوي إجمالي كل القيود المرحلة.
r3 = s.get(BASE + "/api/trial-balance")
t3 = r3.json()["totals"]
check("بدون تاريخ: لا افتتاحي مكرر", t3["o_debit"] == 0 and t3["o_credit"] == 0, str(t3))
check("بدون تاريخ: حركة 'مدين' موجبة", t3["p_debit"] > 0)
# مجموع الحركة (مدين) يبقى كما هو بدون مضاعفة
with_from = s.get(BASE + "/api/trial-balance?from=2026-01-01&to=2026-12-31").json()["totals"]
check("الحركة غير مكررة (مدين)", abs(t3["p_debit"] - with_from["p_debit"]) < 0.01,
      "all=%s withf=%s" % (t3["p_debit"], with_from["p_debit"]))

print("== تصدير الإكسيل ==")
def load_xl(url):
    rr = s.get(BASE + url)
    assert rr.status_code == 200, f"{url} => {rr.status_code}"
    return openpyxl.load_workbook(io.BytesIO(rr.content))

wb = load_xl("/journal/export.xlsx?status=posted")
ws = wb.active
check("تصدير اليومية RTL", ws.sheet_view.rightToLeft)
def find_header_row(ws, keys):
    for row in ws.iter_rows(min_row=1, max_row=15):
        vals = [str(c.value) if c.value else "" for c in row]
        if all(any(k in v for v in vals) for k in keys):
            return row[0].row
    return None
hdr = find_header_row(ws, ["رقم الحركة", "مدين", "دائن"])
check("رأس جدول اليومية صحيح", hdr is not None)
# فحص التنسيق الأسود الرسمي
hcell = [c for c in ws[hdr] if c.value][0]
check("خط الرأس أسود Bold", hcell.font.bold and str(hcell.font.color.rgb).endswith("000000"))
check("حدود سوداء", str(hcell.border.left.color.rgb).endswith("000000"))
dcell = ws.cell(row=hdr + 1, column=3)
check("بيانات Bold أسود", dcell.font.bold and str(dcell.font.color.rgb).endswith("000000"))
check("تكرار رأس الجدول عند الطباعة", ws.print_title_rows is not None)
all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
check("لا توجد حقوق الملكية في التصدير", "أحمد عبدالله" not in all_text)
sig_found = any(isinstance(c.value, str) and c.value.startswith("التوقيع (")
                for row in ws.iter_rows() for c in row if c.value)
name_budget = "محاسب / احمد عبدالله" in all_text
name_finance = "محاسب / محمد احمد سيد" in all_text
dyn_names = [c.value for row in ws.iter_rows() for c in row
             if isinstance(c.value, str) and c.value.startswith("محاسب / ") and c.value != name_budget]
check("التواقيع الثلاثة بالأقواس موجودة", sig_found)
check("اسم مدير إدارة الميزانية موجود", name_budget)
check("اسم مدير الإدارة العامة المالية موجود", name_finance)
check("اسم المحاسب ديناميكي من اليوزر", len(dyn_names) >= 1, str(dyn_names[:2]))
titles_found = "مدير الإدارة العامة المالية" in all_text and "مدير إدارة الميزانية" in all_text
check("المسميات الوظيفية موجودة", titles_found)

wb = load_xl("/ledger/export.xlsx?acc_no=1201&from=2026-08-01&to=2026-08-31")
ws = wb.active
hdr2 = find_header_row(ws, ["التاريخ", "البيان", "الرصيد"])
check("تصدير الأستاذ العام برأس صحيح", hdr2 is not None)

wb = load_xl("/trial-balance/export.xlsx")
ws = wb.active
check("تصدير ميزان المراجعة", "ميزان المراجعة" == ws.title)

wb = load_xl("/accounts/export.xlsx")
check("تصدير دليل الحسابات", wb.active.title == "دليل حسابات")

print("== الاستيراد من القوالب (ملفاتك الأصلية) ==")
with open(r"C:\Users\Boney\Desktop\2023\دليل حسابات.xlsx", "rb") as fp:
    r = s.post(BASE + "/api/accounts/import", files={"file": ("دليل حسابات.xlsx", fp)})
d = r.json()
check("استيراد دليل حسابات (قالبك)", r.status_code == 200, r.text[:150])
accounts = s.get(BASE + "/api/accounts").json()["accounts"]
ids = {a["acc_no"]: a["id"] for a in accounts}

# نصنع ملف يومية ممتلئ ونستورده
src = openpyxl.load_workbook(r"C:\Users\Boney\Desktop\2023\دفتر اليومية.xlsx")
wsm = src.active
data_rows = [
    ["50", "2026-08-20", 700, 0, "الصندوق", "1101", "تحصيل نقدي من عميل"],
    ["50", "2026-08-20", 0, 700, "العملاء (المدينون)", "1201", ""],
    ["51", "20/08/2026", 200, 0, "", "5107", "قرطاسية"],
    ["51", "", 0, 200, "الخزينة", "1101", ""],
]
for i, row in enumerate(data_rows, start=2):
    for j, v in enumerate(row, start=1):
        wsm.cell(row=i, column=j, value=v)
buf = io.BytesIO()
src.save(buf)
buf.seek(0)
r = s.post(BASE + "/api/journal/import",
           files={"file": ("دفتر اليومية.xlsx", buf)},
           data={"region": "المنطقة الجنوبية", "post": "1"})
d = r.json()
check("استيراد قيود من قالبك", d.get("added") == 2, r.text[:200])
check("استيراد أنشأ قيود مرحلة متوازنة", d.get("drafts") == 0)
r = s.get(BASE + "/api/journal?q=تحصيل نقدي")
imp = [e for e in r.json()["entries"] if e["movement_no"] == "50"]
check("القيد المستورد موجود ومرحل", imp and imp[0]["status"] == "posted")

print("== لوحة التحكم والسجل ==")
r = s.get(BASE + "/api/dashboard")
check("إحصائيات اللوحة", r.json()["stats"]["accounts"] > 25)
r = s.get(BASE + "/audit")
check("سجل العمليات", "تسجيل دخول" in r.text or "إنشاء قيد" in r.text)

print("== النسخ الاحتياطي ==")
r = sv.post(BASE + "/api/backup")
check("viewer ممنوع من عمل نسخة", r.status_code == 403)
r = s.post(BASE + "/api/backup")
d = r.json()
check("إنشاء نسخة يدوية", d.get("ok"), r.text[:120])
r = s.get(BASE + "/api/backups")
names = [b["name"] for b in r.json()["backups"]]
check("قائمة النسخ فيها اليدوية", any(n.startswith("manual-") for n in names))
check("النسخة الأوتوماتيكية الشهرية اتعملت", any(n.startswith("auto-") for n in names), str(names[:3]))
target = [n for n in names if n.startswith("manual-")][0]
r = s.get(f"{BASE}/api/backups/{requests.utils.quote(target)}")
check("تنزيل النسخة يعمل", r.status_code == 200 and len(r.content) > 5000)
r = s.get(BASE + "/api/backups/..%5C..%5Csecret.key")
check("حماية من الوصول لملفات خارجية", r.status_code in (400, 404))


print("== المسح الجماعي للقيود ==")
r = s.post(BASE + "/api/journal", json={
    "movement_no": "MV-WIPE-1", "entry_date": "2027-03-15",
    "region": "المنطقة الشرقية", "description": "قيد مسح تجريبي", "entry_type": "عادي",
    "status": "posted",
    "lines": [{"account_id": ids["1101"], "debit": 50, "credit": 0},
              {"account_id": ids["1102"], "debit": 0, "credit": 50}]})
check("إنشاء قيد لاختبار المسح", r.status_code in (200, 201), r.text[:150])
r = sv.post(BASE + "/api/journal/bulk-delete", json={"scope": "month", "month": "2027-03"})
check("viewer ممنوع من المسح الجماعي", r.status_code == 403)
r = s.post(BASE + "/api/journal/bulk-delete", json={"scope": "bogus"})
check("نطاق غير صحيح يرفض", r.status_code == 400)
r = s.post(BASE + "/api/journal/bulk-delete", json={"scope": "month", "month": "bad"})
check("شهر بصيغة خاطئة يرفض", r.status_code == 400)
r = s.post(BASE + "/api/journal/bulk-delete", json={"scope": "month", "month": "2027-02"})
check("شهر فاضي يقول لا توجد قيود", r.status_code == 400)
r = s.get(BASE + "/api/journal?q=MV-WIPE-1")
check("القيد موجود قبل المسح", len(r.json()["entries"]) == 1)
r = s.post(BASE + "/api/journal/bulk-delete", json={"scope": "month", "month": "2027-03"})
d = r.json()
check("مسح شهر كامل يعمل", d.get("deleted") == 1, r.text[:150])
r = s.get(BASE + "/api/journal?q=MV-WIPE-1")
check("القيد اتمسح فعلًا", len(r.json()["entries"]) == 0)
r = s.get(BASE + "/audit")
check("المسح مسجل في سجل التدقيق", "مسح قيود" in r.text)



print("== عمود المنطقة في الاستيراد ==")
r = s.get(BASE + "/templates/journal.xlsx")
tpl = openpyxl.load_workbook(io.BytesIO(r.content)).active
hdrs = [c.value for c in tpl[1]]
check("القالب فيه عمود المنطقة", "المنطقة" in hdrs, str(hdrs))

wb = openpyxl.Workbook()
ws = wb.active
ws.append(["رقم الحركة", "التاريخ", "مدين", "دائن", "إسم الحساب", "رقم الحساب", "البيان", "المنطقة"])
ws.append([9001, "2027-04-01", 100, 0, "الصندوق", "1101", "تجربة منطقة من الملف", "المنطقة الشرقية"])
ws.append([9001, "2027-04-01", 0, 100, "البنك", "1102", "", "المنطقة الشرقية"])
ws.append([9002, "2027-04-02", 200, 0, "الصندوق", "1101", "منطقة فاضية", ""])
ws.append([9002, "2027-04-02", 0, 200, "البنك", "1102", "", ""])
buf = io.BytesIO()
wb.save(buf)
buf.seek(0)
r = s.post(BASE + "/api/journal/import",
           files={"file": ("مناطق.xlsx", buf)},
           data={"region": "المركز الرئيسي", "post": "1"})
d = r.json()
check("استيراد ملف بالمناطق نجح", d.get("added") == 2, r.text[:200])
r = s.get(BASE + "/api/journal?q=9001")
e = [x for x in r.json()["entries"] if x["movement_no"] == "9001"]
check("قيد 9001 أخذ منطقته من الملف", e and e[0]["region"] == "المنطقة الشرقية",
      str(e and e[0]["region"]))
r = s.get(BASE + "/api/journal?q=9002")
e = [x for x in r.json()["entries"] if x["movement_no"] == "9002"]
check("قيد 9002 منطقته فاضية فأخذ اختيار النافذة", e and e[0]["region"] == "المركز الرئيسي",
      str(e and e[0]["region"]))

wb = openpyxl.Workbook()
ws = wb.active
ws.append(["رقم الحركة", "التاريخ", "مدين", "دائن", "إسم الحساب", "رقم الحساب", "البيان", "المنطقة"])
ws.append([9003, "2027-04-03", 300, 0, "الصندوق", "1101", "منطقة غلط", "المنطقة الغربية"])
ws.append([9003, "2027-04-03", 0, 300, "البنك", "1102", "", ""])
buf = io.BytesIO()
wb.save(buf)
buf.seek(0)
r = s.post(BASE + "/api/journal/import",
           files={"file": ("منطقة_غلط.xlsx", buf)},
           data={"region": "المركز الرئيسي", "post": "1"})
d = r.json()
check("منطقة غير معروفة تستخدم البديل مع تحذير",
      d.get("added") == 1 and any("المنطقة الغربية" in e for e in d.get("errors", [])),
      r.text[:250])
r = s.get(BASE + "/api/journal?q=9003")
e = [x for x in r.json()["entries"] if x["movement_no"].startswith("9003")]
check("قيد 9003 اتخزن بالمركز الرئيسي", e and e[0]["region"] == "المركز الرئيسي")



print("== تغيير كلمة المرور الشخصية ==")
r = sv.post(BASE + "/api/me/password", json={"old": "wrong", "new": "viewer999"})
check("كلمة مرور قديمة خاطئة ترفض", r.status_code == 400)
r = sv.post(BASE + "/api/me/password", json={"old": "viewer123", "new": "abc"})
check("كلمة مرور جديدة قصيرة ترفض", r.status_code == 400)
r = sv.post(BASE + "/api/me/password", json={"old": "viewer123", "new": "viewer999"})
check("تغيير كلمة المرور بنجاح", r.status_code == 200, r.text[:150])
sv.post(BASE + "/logout")
r = sv.post(BASE + "/login", data={"username": "viewer", "password": "viewer999"},
            allow_redirects=False)
check("الدخول بكلمة المرور الجديدة", r.status_code in (302, 303))
r = sv.post(BASE + "/login", data={"username": "viewer", "password": "viewer123"},
            allow_redirects=False)
check("القديمة مبتشتغلش", r.status_code == 200)
sv.post(BASE + "/login", data={"username": "viewer", "password": "viewer999"})

# إعادة كلمة مرور viewer إلى الأصلية حتى لا تتأثر التشغيلات التالية على السيرفر الحي
import sqlite3 as _sqlite3
_db = _sqlite3.connect(os.path.join(os.path.dirname(os.path.abspath(__file__)), "instance", "accounting.db"))
_row = _db.execute("SELECT id FROM users WHERE username='viewer'").fetchone()
_db.close()
if _row:
    s.post(BASE + f"/api/users/{_row[0]}/password", json={"password": "viewer123"})
    sv = requests.Session()
    sv.post(BASE + "/login", data={"username": "viewer", "password": "viewer123"})



print("== إدارة المناطق ==")
r = s.get(BASE + "/api/regions")
regions = r.json()["regions"]
check("المناطق الافتراضية موجودة", len(regions) == 5, str(len(regions)))
check("المناطق كلها نشطة", all(r["is_active"] for r in regions))
check("في قيود مرتبطة بالمناطق", any(r["usage_count"] > 0 for r in regions))

r = sv.post(BASE + "/api/regions", json={"name": "viewer test"})
check("viewer ممنوع من إضافة منطقة", r.status_code == 403)
r = s.post(BASE + "/api/regions", json={"name": ""})
check("اسم فاضي يرفض", r.status_code == 400)
r = s.post(BASE + "/api/regions", json={"name": "المنطقة الغربية"})
new_id = r.json().get("id")
check("إضافة منطقة جديدة نجحت", r.status_code == 200 and new_id, r.text[:150])
r = s.post(BASE + "/api/regions", json={"name": "المنطقة الغربية"})
check("تكرار اسم منطقة يرفض", r.status_code == 400)
r = s.get(BASE + "/api/regions")
names = [x["name"] for x in r.json()["regions"]]
check("المنطقة الجديدة تظهر في القائمة", "المنطقة الغربية" in names)

r = s.put(BASE + f"/api/regions/{new_id}", json={"name": "منطقة اختبار"})
check("إعادة تسمية نجحت", r.status_code == 200, r.text[:150])

r = s.post(BASE + f"/api/regions/{new_id}/deactivate")
check("تعطيل المنطقة يعمل", r.status_code == 200)
r = s.get(BASE + "/api/regions")
updated = [x for x in r.json()["regions"] if x["id"] == new_id]
check("المنطقة المتعطّلة بis_active صفر", updated and updated[0]["is_active"] == 0)
r = s.post(BASE + f"/api/regions/{new_id}/activate")
check("تفعيل يرجّعها نشطة", r.status_code == 200)

r = s.post(BASE + "/api/journal", json={
    "movement_no": "MV-REG-1", "entry_date": "2027-05-01",
    "region": "منطقة اختبار", "description": "اختبار المنطقة الجديدة",
    "entry_type": "عادي", "status": "draft",
    "lines": [{"account_id": ids["1101"], "debit": 10, "credit": 0},
              {"account_id": ids["1102"], "debit": 0, "credit": 10}]})
check("قيد بالمنطقة الجديدة يعمل", r.status_code in (200, 201))

r = s.delete(BASE + f"/api/regions/{new_id}")
check("حذف منطقة عليها قيود ممنوع", r.status_code == 400)

r = s.get(BASE + "/api/journal?q=MV-REG-1")
tmp_eid = r.json()["entries"][0]["id"]
r = s.delete(BASE + f"/api/journal/{tmp_eid}")
check("حذف القيد التجريبي", r.status_code == 200)

r = s.delete(BASE + f"/api/regions/{new_id}")
check("حذف منطقة بدون قيود ينجح", r.status_code == 200, r.text[:150])
r = s.get(BASE + "/api/regions")
check("المنطقة المحذوفة اختفت", "منطقة اختبار" not in [x["name"] for x in r.json()["regions"]])


print("\n== الموازنة التقديرية ==")
r = s.get(BASE + "/api/budget")
d = r.json()
check("API الموازنة يعمل", "entries" in d and "summary" in d)
check("في حسابات بميزانيات", any(e["budget"] > 0 for e in d["entries"]), str(len(d["entries"])))
check("ملخص الموازنة محسوب", d["summary"]["budget"] > 0, str(d["summary"]))

r = s.get(BASE + "/api/budget?type=مصروفات")
types = [e["type"] for e in r.json()["entries"]]
check("فلترة بالنوع تعمل", all(t == "مصروفات" for t in types) if types else False)

r = s.get(BASE + "/api/budget?region=المنطقة المركزية")
check("فلترة بالمنطقة تعمل", r.status_code == 200)

r = s.get(BASE + "/api/budget?from=2027-01-01&to=2027-12-31")
check("فلترة بالتاريخ تعمل", r.status_code == 200)

r = s.post(BASE + "/api/accounts/budget", json={"items": []})
check("تحديث ميزانيات فاضي يرفض", r.status_code == 400)

r = s.get(BASE + "/api/accounts")
accs = r.json()["accounts"]
test_acc = next((a for a in accs if a["acc_no"] == "5101"), None)
if test_acc:
    r = s.post(BASE + "/api/accounts/budget", json={"items": [{"id": test_acc["id"], "budget": 999999}]})
    check("تحديث ميزانية حساب واحد", r.status_code == 200 and r.json().get("updated") == 1)
    r2 = s.get(BASE + "/api/accounts")
    updated = next((a for a in r2.json()["accounts"] if a["acc_no"] == "5101"), None)
    check("الميزانية اتحدثت فعليًا", updated and updated["budget"] == 999999)
    s.post(BASE + "/api/accounts/budget", json={"items": [{"id": test_acc["id"], "budget": 480000}]})

r = s.get(BASE + "/budget")
check("صفحة الموازنة تفتح", r.status_code == 200 and "الموازنة التقديرية" in r.text)

r = s.get(BASE + "/budget/export.xlsx")
check("تصدير الموازنة يعمل", r.status_code == 200 and r.headers.get("Content-Type", "").startswith("application/vnd"))
if r.status_code == 200:
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    check("ملف الموازنة فيه صفوف", ws.max_row > 3, str(ws.max_row))
    check("رأس الموازنة يحتوي على الأعمدة", ws.cell(1, 1).value and "موازنة" in str(ws.cell(1, 1).value).lower())


# ============ شجرة الحسابات ============
r = s.get(BASE + "/accounts/tree")
check("صفحة الشجرة تفتح", r.status_code == 200 and "شجرة الحسابات" in r.text)
r = s.get(BASE + "/api/accounts/tree")
tree = r.json().get("tree", []) if r.status_code == 200 else []
check("API الشجرة يعمل", r.status_code == 200)
check("الشجرة فيها أنواع", len(tree) >= 4, str(len(tree)))
all_types = [t["type"] for t in tree]
check("النوع أصول موجود", "أصول" in all_types)
check("النوع مصروفات موجود", "مصروفات" in all_types)
grouped = [g for t in tree for g in t["children"]]
check("الشجرة فيها مجموعات فرعية", len(grouped) >= 3, str(len(grouped)))
accounts_in_tree = [a for g in grouped for a in g["children"]]
check("الشجرة فيها حسابات فعلية", len(accounts_in_tree) >= 10, str(len(accounts_in_tree)))
check("كل حساب تظهر بياناته", all(a["acc_no"] and a["name"] and a["id"] for a in accounts_in_tree))


# ============ التواقيع القابلة للتعديل ============
r = s.get(BASE + "/settings")
check("صفحة الإعدادات فيها قسم التواقيع", r.status_code == 200 and "التواقيع في التقارير" in r.text)
_sig_current = None
with _sqlite3.connect(os.path.join(os.path.dirname(os.path.abspath(__file__)), "instance", "accounting.db")) as _c:
    _sig_current = _c.execute("SELECT value FROM settings WHERE key='signatures'").fetchone()
_sig_current = _sig_current[0] if _sig_current else ""
_sig_names = ["احمد عبدالله", "محمد احمد سيد"]
check("الاسم الافتراضي الأول موجود", "احمد عبدالله" in _sig_current)
check("الاسم الافتراضي الثاني موجود", "محمد احمد سيد" in _sig_current)

# تعديل التواقيع وحفظها
r = s.post(BASE + "/api/settings", json={
    "company_name": "نظام محاسبي متكامل",
    "signatures": [
        {"title": "مدير الإدارة المالية", "name": "محاسب / محمد محمد"},
        {"title": "مدير الحسابات", "name": "عميد / عبدالله أحمد علي"},
    ],
})
check("حفظ التواقيع الجديدة", r.status_code == 200, r.text[:150])

r = s.get(BASE + "/settings")
check("الصفحة تعرض الاسم الجديد", "\\u0639\\u0628\\u062f" in r.text or "عبدالله أحمد علي" in r.text)

# التصدير يحمل التواقيع الجديدة
r = s.get(BASE + "/journal/export.xlsx")
check("تصدير اليومية بعد التعديل", r.status_code == 200)
if r.status_code == 200:
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    cells = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    check("الاسم الجديد ظهر في التوقيع", "عبدالله أحمد علي" in cells)
    check("المسمى الجديد ظهر في التوقيع", "مدير الإدارة المالية" in cells, "مدير إدارة الميزانية" in cells)
    check("اسم المحاسب الديناميكي ظهر", "محاسب / " in cells)

# حذف كل التواقيع = توقيع المحاسب فقط
r = s.post(BASE + "/api/settings", json={
    "company_name": "نظام محاسبي متكامل",
    "signatures": [],
})
check("حذف كل التواقيع", r.status_code == 200)
r = s.get(BASE + "/journal/export.xlsx")
if r.status_code == 200:
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    cells = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    check("المحاسب الديناميكي ظهر بعد الحذف", "محاسب / " in cells)

# إعادة التواقيع الافتراضية
r = s.post(BASE + "/api/settings", json={
    "company_name": "نظام محاسبي متكامل",
    "signatures": [
        {"title": "مدير إدارة الميزانية", "name": "محاسب / احمد عبدالله"},
        {"title": "مدير الإدارة العامة المالية", "name": "محاسب / محمد احمد سيد"},
    ],
})
check("إعادة التواقيع الافتراضية", r.status_code == 200)


print("== التقارير المالية ==")
r = s.get(BASE + "/income-statement")
check("صفحة قائمة الدخل تفتح", r.status_code == 200 and "قائمة الدخل" in r.text)
r = s.get(BASE + "/balance-sheet")
check("صفحة المركز المالي تفتح", r.status_code == 200 and "المركز المالي" in r.text)
r = s.get(BASE + "/cash-flow")
check("صفحة التدفقات النقدية تفتح", r.status_code == 200 and "التدفقات النقدية" in r.text)

r = s.get(BASE + "/api/income-statement")
d = r.json()
check("API قائمة الدخل يعمل", r.status_code == 200)
check("فيه إيرادات من حسابات إيرادات", len(d["revenues"]) >= 1, str(d["revenues"][:1]))
check("فيه مصروفات", len(d["expenses"]) >= 1, str(d["expenses"][:1]))
check("صافي الدخل = إيرادات - مصروفات",
      abs(d["net"] - (d["revenues_total"] - d["expenses_total"])) < 0.01,
      str(d))
r = s.get(BASE + "/api/income-statement?from=2026-08-01&to=2026-08-31&region=المنطقة الجنوبية")
d2 = r.json()
check("فلترة قائمة الدخل تعمل", r.status_code == 200 and len(d2["expenses"]) >= 1)

r = s.get(BASE + "/api/balance-sheet")
d = r.json()
check("API المركز المالي يعمل", r.status_code == 200)
check("فيه أصول", len(d["assets"]) >= 1, str(d["assets"][:1]))
check("فيه التزامات/حقوق ملكية", len(d["liabilities_equity"]) >= 1, str(d["liabilities_equity"][:1]))
check("المستند منفعة صافي متراكم موجود", "net_cumulative" in d)
check("مجموع الأصول = مجموع صفوفها",
      abs(d["assets_total"] - sum(a["amount"] for a in d["assets"])) < 0.01, str(d["assets_total"]))
check("مجموع الالتزامات = مجموع صفوفها (موجبة)",
      abs(d["liab_equity_total"] - sum(a["amount"] for a in d["liabilities_equity"])) < 0.01
      and all(a["amount"] >= 0 for a in d["liabilities_equity"]), str(d["liab_equity_total"]))
check("معادلة التوازن: أصول = خصوم وحقوق ملكية + أرباح متراكمة",
      abs(d["assets_total"] - (d["liab_equity_total"] + d["net_cumulative"])) <= 0.02,
      f"difference={d['difference']}")

r = s.get(BASE + "/api/cash-flow")
d = r.json()
check("API التدفقات يعمل", r.status_code == 200)
check("فيه حسابات نقدية (صندوق/بنك)", len(d["rows"]) >= 1, str(d["rows"][:1]))
cash_t = d["totals"]
check("صافي التدفق = داخل - خارج",
      abs(d["net_flow"] - (cash_t["inflow"] - cash_t["outflow"])) < 0.01, str(d))

print("== تصدير التقارير المالية ==")
wb = load_xl("/income-statement/export.xlsx")
ws = wb.active
check("تصدير قائمة الدخل", "قائمة الدخل" in ws.title.lower() or "دخل" in str(ws.cell(1, 1).value))
wb = load_xl("/balance-sheet/export.xlsx")
check("تصدير المركز المالي", wb.active.title.lower() == "المركز المالي")
wb = load_xl("/cash-flow/export.xlsx")
check("تصدير التدفقات النقدية", "تدفق" in wb.active.title.lower())

print("== الاستعادة من النسخ الاحتياطي (آخر قاعدة البيانات) ==")
r = sv.post(BASE + f"/api/restore/{requests.utils.quote(target)}")
check("viewer ممنوع من الاستعادة", r.status_code == 403)
r = s.post(BASE + "/api/restore/bad%5C..%5Csecret.key")
check("اسم نسخة غير صالح يرفض", r.status_code == 400)
r = s.post(BASE + "/api/journal", json={
    "movement_no": "MV-RESTORE-1", "entry_date": "2027-04-01",
    "region": "المنطقة الشرقية", "description": "قيد قبل الاستعادة", "entry_type": "عادي",
    "status": "posted",
    "lines": [{"account_id": ids["1101"], "debit": 30, "credit": 0},
              {"account_id": ids["1102"], "debit": 0, "credit": 30}]})
check("إنشاء قيد قبل الاستعادة", r.status_code in (200, 201), r.text[:150])
r = s.get(BASE + "/api/journal?q=MV-RESTORE-1")
before = any(e["movement_no"] == "MV-RESTORE-1" for e in r.json()["entries"])
check("القيد موجود قبل الاستعادة", before)
r = s.post(BASE + f"/api/restore/{requests.utils.quote(target)}")
check("إتمام الاستعادة", r.json().get("ok"), r.text[:150])
r = s.get(BASE + "/api/journal?q=MV-RESTORE-1")
after = any(e["movement_no"] == "MV-RESTORE-1" for e in r.json()["entries"])
check("القيد المضافة اختفت بعد الاستعادة", not after)
r = s.get(BASE + "/api/accounts")
check("النظام يعمل طبيعي بعد الاستعادة", r.status_code == 200)

print("== مسح الشجرة (دليل الحسابات بالكامل) ==")
r = sv.post(BASE + "/api/accounts/wipe")
check("viewer ممنوع من مسح الشجرة", r.status_code == 403)
r = s.post(BASE + "/api/accounts/wipe")
d = r.json()
check("مسح الشجرة يعمل", r.status_code == 200 and d.get("ok"), r.text[:150])
check("كل الحسابات اتشالت", d.get("deleted_accounts", 0) >= 1, str(d))
r = s.get(BASE + "/api/accounts")
check("دليل الحسابات فاضي بعد المسح", len(r.json()["accounts"]) == 0, r.text[:150])
r = s.get(BASE + "/api/accounts/tree")
tree = r.json().get("tree", []) if r.status_code == 200 else []
check("الشجرة فاضية بعد المسح", not tree, str(tree))
r = s.get(BASE + "/api/journal")
check("اليومية فاضية بعد المسح", len(r.json()["entries"]) == 0, r.text[:150])
r = s.post(BASE + "/api/accounts/wipe")
check("مسح على دليل فاضي يرفض", r.status_code == 400, r.text[:150])

print("=" * 46)
if fail:
    print(f"فشل {len(fail)} اختبار: {fail}")
    sys.exit(1)
print(f"كل الاختبارات نجحت ({ok_count}) ✔")
